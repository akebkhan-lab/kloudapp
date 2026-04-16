import io
import re

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="POS + Foodpanda Sales Combiner", layout="wide")

REQUIRED_OUTPUT_COLS = ["Item Name", "POS Qty", "Foodpanda Qty", "Total Qty"]

OUTLETS = [
    "Peyari Tehari Agrabad",
    "Peyari Tehari Lalkhan Bazaar",
    "Peyari Tehari Chawkbazar",
    "Peyari Tehari Gulshan",
]


def clean_text(value: str) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value: str) -> str:
    text = clean_text(value).lower()
    replacements = {
        "1 kg": "1kg",
        "1  kg": "1kg",
        "1 liter": "1l",
        "1 litre": "1l",
        "250 ml": "250ml",
        "500 ml": "500ml",
        "full plate": "full",
        "half plate": "half",
        "matka": "matka",
        "teheri": "tehari",
        "shorbot": "sorbot",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = text.replace("-", " ")
    text = re.sub(r"[^a-z0-9 ]", "", text)
    tokens = [t for t in text.split() if t]

    noise = {"normal", "pcs", "pc"}
    tokens = [t for t in tokens if t not in noise]

    return " ".join(tokens)


def parse_pos_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()

    if name.endswith(".xls") or name.endswith(".html"):
        uploaded_file.seek(0)
        tables = pd.read_html(uploaded_file)
        if not tables:
            raise ValueError("Could not read any table from the POS file.")

        raw = tables[0].copy()
        raw.columns = [
            " | ".join([str(x) for x in col if str(x) != "nan"]).strip()
            if isinstance(col, tuple)
            else str(col)
            for col in raw.columns
        ]

        raw = raw.reset_index(drop=True)
        if raw.shape[1] < 9:
            raise ValueError("POS table shape was not recognized.")

        raw.columns = [
            "Department Name",
            "Group Name",
            "Item Name",
            "Portion Name",
            "Price",
            "Quantity",
            "Net Amount",
            "Gross",
            "Sub Total",
        ]

        item_df = raw[raw["Department Name"].astype(str).str.strip().eq("All")].copy()
    else:
        uploaded_file.seek(0)
        item_df = pd.read_excel(uploaded_file)
        item_df.columns = [clean_text(c) for c in item_df.columns]

    rename_map = {}
    for col in item_df.columns:
        low = clean_text(col).lower()
        if low == "item name":
            rename_map[col] = "Item Name"
        elif low == "price":
            rename_map[col] = "Price"
        elif low == "quantity":
            rename_map[col] = "Quantity"
        elif low == "net amount":
            rename_map[col] = "Net Amount"
    item_df = item_df.rename(columns=rename_map)

    required = ["Item Name", "Price", "Quantity", "Net Amount"]
    missing = [c for c in required if c not in item_df.columns]
    if missing:
        raise ValueError(f"POS file is missing required columns: {missing}")

    item_df = item_df[required].copy()
    item_df["Item Name"] = item_df["Item Name"].map(clean_text)
    item_df = item_df[item_df["Item Name"] != ""]
    item_df = item_df[~item_df["Item Name"].str.lower().str.contains("total|powered by", na=False)]

    for c in ["Price", "Quantity", "Net Amount"]:
        item_df[c] = pd.to_numeric(item_df[c], errors="coerce")

    item_df = item_df.dropna(subset=["Quantity"])
    return item_df.reset_index(drop=True)


def parse_foodpanda_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)

    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    df.columns = [clean_text(c) for c in df.columns]

    dish_col = None
    qty_col = None
    sales_col = None

    for col in df.columns:
        low = col.lower()
        if low in {"dish", "dish name", "item name", "product"}:
            dish_col = col
        elif low in {"total", "quantity", "qty", "items sold"}:
            qty_col = col
        elif low in {"sales", "net sales", "amount"}:
            sales_col = col

    if not dish_col or not qty_col:
        raise ValueError(
            "Foodpanda file was not recognized. Need at least a dish column and a quantity/total column."
        )

    out = df[[dish_col, qty_col] + ([sales_col] if sales_col else [])].copy()
    out = out.rename(columns={dish_col: "Item Name", qty_col: "Quantity"})

    if sales_col:
        out = out.rename(columns={sales_col: "Sales"})
    else:
        out["Sales"] = pd.NA

    out["Item Name"] = out["Item Name"].map(clean_text)
    out["Quantity"] = pd.to_numeric(out["Quantity"], errors="coerce")

    if "Sales" in out.columns:
        out["Sales"] = out["Sales"].astype(str).str.replace(",", "", regex=False)
        out["Sales"] = pd.to_numeric(out["Sales"], errors="coerce")

    out = out.dropna(subset=["Item Name", "Quantity"])
    out = out[out["Item Name"] != ""]
    return out.reset_index(drop=True)


def build_mapping_df(pos_df: pd.DataFrame, fp_df: pd.DataFrame) -> pd.DataFrame:
    pos_names = sorted(pos_df["Item Name"].dropna().astype(str).unique().tolist())
    fp_names = sorted(fp_df["Item Name"].dropna().astype(str).unique().tolist())

    fp_key_map = {}
    for name in fp_names:
        key = normalize_key(name)
        fp_key_map.setdefault(key, []).append(name)

    rows = []
    used_fp = set()

    for pos_name in pos_names:
        key = normalize_key(pos_name)
        candidates = fp_key_map.get(key, [])
        suggested = candidates[0] if candidates else ""
        if suggested:
            used_fp.add(suggested)

        rows.append(
            {
                "POS Item Name": pos_name,
                "Suggested Foodpanda Match": suggested,
                "Final Standard Name": pos_name,
            }
        )

    for fp_name in fp_names:
        if fp_name not in used_fp:
            rows.append(
                {
                    "POS Item Name": "",
                    "Suggested Foodpanda Match": fp_name,
                    "Final Standard Name": fp_name,
                }
            )

    return pd.DataFrame(rows)


def apply_mapping(
    pos_df: pd.DataFrame,
    fp_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    mapping_df = mapping_df.copy()
    mapping_df.columns = [clean_text(c) for c in mapping_df.columns]

    needed = {"POS Item Name", "Suggested Foodpanda Match", "Final Standard Name"}
    if not needed.issubset(set(mapping_df.columns)):
        raise ValueError(
            "Mapping file must contain columns: POS Item Name, Suggested Foodpanda Match, Final Standard Name"
        )

    mapping_df = mapping_df.fillna("")
    pos_to_final = {}
    fp_to_final = {}

    for _, row in mapping_df.iterrows():
        pos_item = clean_text(row["POS Item Name"])
        fp_item = clean_text(row["Suggested Foodpanda Match"])
        final_name = clean_text(row["Final Standard Name"]) or pos_item or fp_item

        if pos_item:
            pos_to_final[pos_item] = final_name
        if fp_item:
            fp_to_final[fp_item] = final_name

    pos_key_to_final = {
        normalize_key(k): v for k, v in pos_to_final.items() if clean_text(k)
    }
    fp_key_to_final = {
        normalize_key(k): v for k, v in fp_to_final.items() if clean_text(k)
    }

    pos_clean = pos_df.copy()
    pos_clean["Standard Item Name"] = pos_clean["Item Name"].map(
        lambda x: pos_to_final.get(
            clean_text(x),
            pos_key_to_final.get(normalize_key(x), clean_text(x))
        )
    )

    fp_clean = fp_df.copy()
    fp_clean["Standard Item Name"] = fp_clean["Item Name"].map(
        lambda x: fp_to_final.get(
            clean_text(x),
            pos_key_to_final.get(
                normalize_key(x),
                fp_key_to_final.get(normalize_key(x), clean_text(x))
            )
        )
    )

    mapped_final_names = set(mapping_df["Final Standard Name"].map(clean_text))
    fp_unmatched = fp_clean[~fp_clean["Standard Item Name"].isin(mapped_final_names)].copy()

    return pos_clean, fp_clean, fp_unmatched


def build_outlet_summaries(pos_clean: pd.DataFrame, fp_clean: pd.DataFrame) -> pd.DataFrame:
    outlet_names = sorted(
        set(pos_clean["Outlet"].dropna().unique()).union(set(fp_clean["Outlet"].dropna().unique()))
    )

    results = []

    for outlet in outlet_names:
        pos_outlet = pos_clean[pos_clean["Outlet"] == outlet].copy()
        fp_outlet = fp_clean[fp_clean["Outlet"] == outlet].copy()

        pos_sum = (
            pos_outlet.groupby("Standard Item Name", as_index=False)["Quantity"]
            .sum()
            .rename(columns={"Standard Item Name": "Item Name", "Quantity": "POS Qty"})
        )

        fp_sum = (
            fp_outlet.groupby("Standard Item Name", as_index=False)["Quantity"]
            .sum()
            .rename(columns={"Standard Item Name": "Item Name", "Quantity": "Foodpanda Qty"})
        )

        final = pos_sum.merge(fp_sum, on="Item Name", how="outer")
        if final.empty:
            continue

        final[["POS Qty", "Foodpanda Qty"]] = final[["POS Qty", "Foodpanda Qty"]].fillna(0)
        final["Total Qty"] = final["POS Qty"] + final["Foodpanda Qty"]

        for col in ["POS Qty", "Foodpanda Qty", "Total Qty"]:
            final[col] = final[col].astype(int)

        final["Outlet"] = outlet
        final = final[["Outlet", "Item Name", "POS Qty", "Foodpanda Qty", "Total Qty"]]
        final = final.sort_values(["Total Qty", "Item Name"], ascending=[False, True]).reset_index(drop=True)
        results.append(final)

    if not results:
        return pd.DataFrame(columns=["Outlet", "Item Name", "POS Qty", "Foodpanda Qty", "Total Qty"])

    return pd.concat(results, ignore_index=True)


def build_grand_summary(outlet_summary_df: pd.DataFrame) -> pd.DataFrame:
    if outlet_summary_df.empty:
        return pd.DataFrame(columns=REQUIRED_OUTPUT_COLS)

    grand = (
        outlet_summary_df.groupby("Item Name", as_index=False)[["POS Qty", "Foodpanda Qty", "Total Qty"]]
        .sum()
        .sort_values(["Total Qty", "Item Name"], ascending=[False, True])
        .reset_index(drop=True)
    )

    for col in ["POS Qty", "Foodpanda Qty", "Total Qty"]:
        grand[col] = grand[col].astype(int)

    return grand[REQUIRED_OUTPUT_COLS]


def build_beef_tehari_equivalent_summary(outlet_summary_df: pd.DataFrame) -> pd.DataFrame:
    if outlet_summary_df.empty:
        return pd.DataFrame(columns=["Outlet", "Equivalent Full Plate Beef Tehari"])

    conversion_map = {
        "1 KG Beef Matka": 2.0,
        "Full Plate Beef Tehari": 1.0,
        "Half Plate Beef Tehari": 0.5,
    }

    beef_df = outlet_summary_df.copy()
    beef_df["Conversion Factor"] = beef_df["Item Name"].map(conversion_map)
    beef_df = beef_df.dropna(subset=["Conversion Factor"]).copy()

    beef_df["Equivalent Full Plate Beef Tehari"] = (
        beef_df["Total Qty"] * beef_df["Conversion Factor"]
    )

    summary = (
        beef_df.groupby("Outlet", as_index=False)["Equivalent Full Plate Beef Tehari"]
        .sum()
        .sort_values("Outlet")
        .reset_index(drop=True)
    )

    return summary


def build_beef_tehari_equivalent_grand_total(beef_equiv_df: pd.DataFrame) -> float:
    if beef_equiv_df.empty:
        return 0.0
    return float(beef_equiv_df["Equivalent Full Plate Beef Tehari"].sum())


def to_excel_bytes(
    grand_summary_df: pd.DataFrame,
    outlet_summary_df: pd.DataFrame,
    beef_equiv_df: pd.DataFrame,
    pos_clean: pd.DataFrame,
    fp_clean: pd.DataFrame,
    mapping_df: pd.DataFrame,
    unmatched_fp: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        grand_summary_df.to_excel(writer, index=False, sheet_name="Grand Summary")
        outlet_summary_df.to_excel(writer, index=False, sheet_name="Outlet Summary")
        beef_equiv_df.to_excel(writer, index=False, sheet_name="Beef Equivalent")
        pos_clean.to_excel(writer, index=False, sheet_name="POS Clean")
        fp_clean.to_excel(writer, index=False, sheet_name="Foodpanda Clean")
        mapping_df.to_excel(writer, index=False, sheet_name="Mapping")
        unmatched_fp.to_excel(writer, index=False, sheet_name="Unmatched FP Items")

        for outlet in OUTLETS:
            outlet_df = outlet_summary_df[outlet_summary_df["Outlet"] == outlet].copy()
            safe_sheet_name = outlet.replace("Peyari Tehari ", "")[:31]
            if not outlet_df.empty:
                outlet_df.to_excel(writer, index=False, sheet_name=safe_sheet_name)

        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            df_lookup = None

            if sheet_name == "Grand Summary":
                df_lookup = grand_summary_df
            elif sheet_name == "Outlet Summary":
                df_lookup = outlet_summary_df
            elif sheet_name == "Beef Equivalent":
                df_lookup = beef_equiv_df
            elif sheet_name == "POS Clean":
                df_lookup = pos_clean
            elif sheet_name == "Foodpanda Clean":
                df_lookup = fp_clean
            elif sheet_name == "Mapping":
                df_lookup = mapping_df
            elif sheet_name == "Unmatched FP Items":
                df_lookup = unmatched_fp
            else:
                original_outlet = next(
                    (o for o in OUTLETS if o.replace("Peyari Tehari ", "")[:31] == sheet_name),
                    None,
                )
                if original_outlet:
                    df_lookup = outlet_summary_df[outlet_summary_df["Outlet"] == original_outlet].copy()

            if df_lookup is None:
                continue

            for idx, col in enumerate(df_lookup.columns, start=1):
                values = [str(col)] + [str(v) for v in df_lookup[col].head(200).tolist()]
                max_len = max(len(v) for v in values) if values else len(str(col))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 35)

    output.seek(0)
    return output.getvalue()


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


st.title("POS + Foodpanda Item Quantity Combiner")
st.caption(
    "Upload POS and Foodpanda files under each outlet, review the mapping, and download outlet-wise plus grand summaries."
)

with st.sidebar:
    st.header("Optional Mapping File")
    mapping_upload = st.file_uploader(
        "Upload saved mapping CSV/XLSX",
        type=["csv", "xlsx", "xls"],
        help="Use this if you already finalized your item mapping and want to reuse it.",
    )

st.subheader("Outlet-wise Uploads")

outlet_uploads = {}
for outlet in OUTLETS:
    with st.expander(outlet, expanded=False):
        pos_files = st.file_uploader(
            f"{outlet} - Upload POS report(s)",
            type=["xls", "xlsx", "html"],
            accept_multiple_files=True,
            key=f"pos_{outlet}",
        )
        fp_files = st.file_uploader(
            f"{outlet} - Upload Foodpanda report(s)",
            type=["csv", "xlsx", "xls"],
            accept_multiple_files=True,
            key=f"fp_{outlet}",
        )
        outlet_uploads[outlet] = {
            "pos_files": pos_files,
            "fp_files": fp_files,
        }

has_any_files = any(
    outlet_uploads[outlet]["pos_files"] or outlet_uploads[outlet]["fp_files"]
    for outlet in OUTLETS
)

if not has_any_files:
    st.info("Upload at least one POS or Foodpanda file under the outlet sections to begin.")
    st.stop()

all_pos = []
all_fp = []
parse_errors = []

for outlet, files in outlet_uploads.items():
    for pos_file in files["pos_files"]:
        try:
            pos_df_single = parse_pos_file(pos_file)
            pos_df_single["Outlet"] = outlet
            pos_df_single["Source File"] = pos_file.name
            all_pos.append(pos_df_single)
        except Exception as exc:
            parse_errors.append(f"{outlet} - POS file '{pos_file.name}': {exc}")

    for fp_file in files["fp_files"]:
        try:
            fp_df_single = parse_foodpanda_file(fp_file)
            fp_df_single["Outlet"] = outlet
            fp_df_single["Source File"] = fp_file.name
            all_fp.append(fp_df_single)
        except Exception as exc:
            parse_errors.append(f"{outlet} - Foodpanda file '{fp_file.name}': {exc}")

if parse_errors:
    for err in parse_errors:
        st.error(err)

if not all_pos and not all_fp:
    st.warning("No valid files could be parsed.")
    st.stop()

if all_pos:
    pos_df = pd.concat(all_pos, ignore_index=True)
else:
    pos_df = pd.DataFrame(columns=["Item Name", "Price", "Quantity", "Net Amount", "Outlet", "Source File"])

if all_fp:
    fp_df = pd.concat(all_fp, ignore_index=True)
else:
    fp_df = pd.DataFrame(columns=["Item Name", "Quantity", "Sales", "Outlet", "Source File"])

st.success("Files parsed successfully.")

m1, m2, m3, m4 = st.columns(4)
m1.metric("POS Rows", len(pos_df))
m2.metric("Foodpanda Rows", len(fp_df))
m3.metric("POS Qty Total", int(pos_df["Quantity"].sum()) if not pos_df.empty else 0)
m4.metric("Foodpanda Qty Total", int(fp_df["Quantity"].sum()) if not fp_df.empty else 0)

with st.expander("Preview cleaned source data", expanded=False):
    a, b = st.columns(2)
    with a:
        st.subheader("POS cleaned")
        st.dataframe(pos_df, use_container_width=True)
    with b:
        st.subheader("Foodpanda cleaned")
        st.dataframe(fp_df, use_container_width=True)

try:
    if mapping_upload is not None:
        if mapping_upload.name.lower().endswith(".csv"):
            mapping_df = pd.read_csv(mapping_upload)
        else:
            mapping_df = pd.read_excel(mapping_upload)
        mapping_df.columns = [clean_text(c) for c in mapping_df.columns]
    else:
        mapping_df = build_mapping_df(pos_df, fp_df)
except Exception as exc:
    st.error(f"Could not create/load mapping table: {exc}")
    st.stop()

st.subheader("Item Mapping")
st.write(
    "Confirm each POS item’s matching Foodpanda item and set the final standard item name. "
    "This mapping will be applied across all outlets."
)

edited_mapping = st.data_editor(
    mapping_df,
    use_container_width=True,
    num_rows="dynamic",
    key="mapping_editor",
)

try:
    pos_clean, fp_clean, unmatched_fp = apply_mapping(pos_df, fp_df, edited_mapping)
    outlet_summary_df = build_outlet_summaries(pos_clean, fp_clean)
    grand_summary_df = build_grand_summary(outlet_summary_df)
    beef_equiv_df = build_beef_tehari_equivalent_summary(outlet_summary_df)
    beef_equiv_grand_total = build_beef_tehari_equivalent_grand_total(beef_equiv_df)
except Exception as exc:
    st.error(f"Mapping/merge error: {exc}")
    st.stop()

st.subheader("Outlet-wise Combined Data")

if outlet_summary_df.empty:
    st.warning("No combined outlet summary could be generated yet.")
else:
    for outlet in OUTLETS:
        outlet_df = outlet_summary_df[outlet_summary_df["Outlet"] == outlet].copy()
        with st.expander(outlet, expanded=False):
            if outlet_df.empty:
                st.info("No combined data for this outlet yet.")
            else:
                st.dataframe(outlet_df, use_container_width=True)

st.subheader("Beef Tehari Equivalent Summary (Full Plate Standard)")
st.caption(
    "Conversion used: 1 KG Beef Matka = 2 Full Plate Beef Tehari, "
    "1 Full Plate Beef Tehari = 1, Half Plate Beef Tehari = 0.5"
)
st.dataframe(beef_equiv_df, use_container_width=True)

b1, b2 = st.columns(2)
with b1:
    st.metric(
        "Total Equivalent Full Plate Beef Tehari",
        f"{beef_equiv_grand_total:,.1f}"
    )
with b2:
    st.metric(
        "Outlets Covered",
        len(beef_equiv_df)
    )

st.subheader("Grand Summary")
st.dataframe(grand_summary_df, use_container_width=True)

s1, s2, s3 = st.columns(3)
s1.metric("Grand POS Qty", int(grand_summary_df["POS Qty"].sum()) if not grand_summary_df.empty else 0)
s2.metric("Grand Foodpanda Qty", int(grand_summary_df["Foodpanda Qty"].sum()) if not grand_summary_df.empty else 0)
s3.metric("Grand Total Qty", int(grand_summary_df["Total Qty"].sum()) if not grand_summary_df.empty else 0)

if len(unmatched_fp) > 0:
    with st.expander(f"Unmatched Foodpanda items ({len(unmatched_fp)})", expanded=True):
        st.warning(
            "These Foodpanda items did not confidently map to a final standard item name from your mapping sheet. "
            "Review them before finalizing."
        )
        cols_to_show = ["Outlet", "Item Name", "Quantity", "Standard Item Name"]
        existing_cols = [c for c in cols_to_show if c in unmatched_fp.columns]
        st.dataframe(unmatched_fp[existing_cols], use_container_width=True)

excel_bytes = to_excel_bytes(
    grand_summary_df=grand_summary_df,
    outlet_summary_df=outlet_summary_df,
    beef_equiv_df=beef_equiv_df,
    pos_clean=pos_clean,
    fp_clean=fp_clean,
    mapping_df=edited_mapping,
    unmatched_fp=unmatched_fp,
)
mapping_csv_bytes = to_csv_bytes(edited_mapping)

c1, c2 = st.columns(2)
with c1:
    st.download_button(
        "Download final workbook (.xlsx)",
        data=excel_bytes,
        file_name="combined_sales_summary.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with c2:
    st.download_button(
        "Download mapping template (.csv)",
        data=mapping_csv_bytes,
        file_name="item_mapping.csv",
        mime="text/csv",
        use_container_width=True,
    )

st.markdown("---")
st.markdown(
    "**How it works:** POS item quantities and Foodpanda item quantities are cleaned outlet-wise, "
    "matched through one shared mapping table, then shown as individual outlet summaries plus one grand summary."
)
